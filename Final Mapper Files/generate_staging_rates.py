#!/usr/bin/env python3
"""Generate staging_rates.js from local equipment and consumables spreadsheets."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EQUIPMENT_XLSX = BASE_DIR / "Schedule_of_Equipment_Rates.xlsx"
CONSUMABLES_XLSX = BASE_DIR / "Consumables.xlsx"
OUTPUT_JS = BASE_DIR / "staging_rates.js"


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "item"


def _read_rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_text(h) for h in rows[0]]
    data: list[dict[str, object]] = []
    for raw in rows[1:]:
        row = {headers[idx]: raw[idx] for idx in range(min(len(headers), len(raw)))}
        if all(v in (None, "") for v in row.values()):
            continue
        data.append(row)
    return data


def build_consumables() -> list[dict[str, object]]:
    rows = _read_rows(CONSUMABLES_XLSX)
    out: list[dict[str, object]] = []
    seen: dict[str, int] = {}
    for row in rows:
        label = _normalize_text(row.get("Supplies / Service"))
        if not label:
            continue
        price = _to_float(row.get("Price"))
        unit = _normalize_text(row.get("Unit of Measurement")) or "Unit"
        base = f"consumable_{_slugify(label)}"
        seen[base] = seen.get(base, 0) + 1
        name = f"{base}_{seen[base]}" if seen[base] > 1 else base
        out.append(
            {
                "name": name,
                "label": label,
                "costRate": round(price, 2),
                "unit": unit,
                "billingModel": "quantity",
                "description": f"{label} ({unit})",
            }
        )
    return out


def build_equipment() -> list[dict[str, object]]:
    rows = _read_rows(EQUIPMENT_XLSX)
    out: list[dict[str, object]] = []
    seen: dict[str, int] = {}
    for row in rows:
        equipment = _normalize_text(row.get("Equipment"))
        manufacturer = _normalize_text(row.get("Manufacturer"))
        spec = _normalize_text(row.get("Specification"))
        capacity = _normalize_text(row.get("Capacity or Size"))
        hp = _normalize_text(row.get("HP"))
        notes = _normalize_text(row.get("Notes"))
        unit = _normalize_text(row.get("Unit")) or "Hour"
        cost_rate = _to_float(row.get("2025 Rates"))
        cost_code = _normalize_text(row.get("Cost Code"))

        if not any([equipment, manufacturer, spec, capacity, hp, notes, cost_code]):
            continue
        if not equipment and not cost_code:
            continue

        key_base = f"equipment_{_slugify(cost_code or equipment)}"
        seen[key_base] = seen.get(key_base, 0) + 1
        name = f"{key_base}_{seen[key_base]}" if seen[key_base] > 1 else key_base

        label_parts = [equipment] if equipment else []
        if manufacturer:
            label_parts.append(manufacturer)
        label = " - ".join(label_parts) if label_parts else (cost_code or "Equipment")

        desc_parts = [p for p in [equipment, manufacturer, capacity] if p]
        description = " | ".join(desc_parts) if desc_parts else label

        out.append(
            {
                "name": name,
                "label": label,
                "costRate": round(cost_rate, 2),
                "unit": unit,
                "billingModel": "hourly" if unit.lower() == "hour" else "quantity",
                "costCode": cost_code,
                "equipment": equipment,
                "manufacturer": manufacturer,
                "specification": spec,
                "capacity": capacity,
                "hp": hp,
                "notes": notes,
                "description": description,
            }
        )
    return out


def main() -> int:
    payload = {
        "consumables": build_consumables(),
        "equipment": build_equipment(),
    }
    OUTPUT_JS.write_text(
        "window.HAZMAT_RATE_CATALOGS = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    print(
        f"Generated {OUTPUT_JS.name}: "
        f"{len(payload['equipment'])} equipment rows, "
        f"{len(payload['consumables'])} consumable rows."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
